import csv
import hashlib
import io
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db import get_db
from app.models import Trade
from app.schemas import TradeImportResponse, TradeImportRow

router = APIRouter(prefix="/v1/trades", tags=["trades"])

REQUIRED_FIELDS = {"account", "symbol", "trade_date", "side", "quantity", "price"}
OPTIONAL_FIELDS = {"fees", "currency", "broker_ref"}


def _parse_date(value: str) -> date:
    candidate = value.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value}")


def _parse_decimal(value: str, field: str) -> Decimal:
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        raise ValueError(f"Invalid decimal value for {field}: {value}") from None


def _normalize_header(value: str) -> str:
    return value.strip().lower()


def _build_trade_uid(row: TradeImportRow) -> str:
    raw = "|".join(
        [
            row.account.strip().upper(),
            row.symbol.strip().upper(),
            row.trade_date.isoformat(),
            row.side.strip().upper(),
            str(row.quantity),
            str(row.price),
            str(row.fees),
            row.currency.strip().upper(),
            (row.broker_ref or "").strip().upper(),
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _row_from_mapping(mapping: dict[str, str], row_number: int) -> TradeImportRow:
    fields = {k: (v if v is not None else "") for k, v in mapping.items()}
    missing = sorted(field for field in REQUIRED_FIELDS if not str(fields.get(field, "")).strip())
    if missing:
        raise ValueError(f"Row {row_number}: missing required fields: {', '.join(missing)}")

    side = str(fields["side"]).strip().upper()
    if side not in {"BUY", "SELL"}:
        raise ValueError(f"Row {row_number}: side must be BUY or SELL")

    quantity = _parse_decimal(str(fields["quantity"]), "quantity")
    price = _parse_decimal(str(fields["price"]), "price")
    fees = _parse_decimal(str(fields.get("fees", "0") or "0"), "fees")

    if quantity <= 0:
        raise ValueError(f"Row {row_number}: quantity must be > 0")
    if price < 0:
        raise ValueError(f"Row {row_number}: price must be >= 0")
    if fees < 0:
        raise ValueError(f"Row {row_number}: fees must be >= 0")

    return TradeImportRow(
        account=str(fields["account"]).strip(),
        symbol=str(fields["symbol"]).strip().upper(),
        trade_date=_parse_date(str(fields["trade_date"])),
        side=side,
        quantity=quantity,
        price=price,
        fees=fees,
        currency=str(fields.get("currency", "USD") or "USD").strip().upper(),
        broker_ref=str(fields.get("broker_ref", "") or "").strip() or None,
    )


def _parse_csv(contents: bytes) -> list[dict[str, str]]:
    stream = io.StringIO(contents.decode("utf-8-sig"))
    reader = csv.DictReader(stream)
    if not reader.fieldnames:
        raise ValueError("CSV file must include headers")
    headers = [_normalize_header(h) for h in reader.fieldnames]

    if not REQUIRED_FIELDS.issubset(set(headers)):
        missing = ", ".join(sorted(REQUIRED_FIELDS.difference(headers)))
        raise ValueError(f"Missing required columns: {missing}")

    rows: list[dict[str, str]] = []
    for row in reader:
        normalized = {_normalize_header(k): v for k, v in row.items() if k is not None}
        rows.append({k: normalized.get(k, "") for k in REQUIRED_FIELDS.union(OPTIONAL_FIELDS)})
    return rows


def _parse_xlsx(contents: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(contents), data_only=True)
    sheet = workbook.active
    raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [_normalize_header(str(h or "")) for h in raw_headers]

    if not REQUIRED_FIELDS.issubset(set(headers)):
        missing = ", ".join(sorted(REQUIRED_FIELDS.difference(headers)))
        raise ValueError(f"Missing required columns: {missing}")

    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        normalized = {headers[idx]: ("" if value is None else str(value)) for idx, value in enumerate(row)}
        if not any(normalized.values()):
            continue
        rows.append({k: normalized.get(k, "") for k in REQUIRED_FIELDS.union(OPTIONAL_FIELDS)})
    return rows


def _parse_upload(filename: str, contents: bytes) -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(contents)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(contents)
    raise ValueError("Only .csv and .xlsx files are supported")


@router.post("/import", response_model=TradeImportResponse)
def import_trades(file: UploadFile = File(...), db: Session = Depends(get_db)) -> TradeImportResponse:
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Filename is required")

    contents = file.file.read()
    if not contents:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        parsed_rows = _parse_upload(file.filename, contents)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(exc)) from exc

    validated_rows: list[TradeImportRow] = []
    for idx, mapping in enumerate(parsed_rows, start=2):
        try:
            validated_rows.append(_row_from_mapping(mapping, row_number=idx))
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_CONTENT, detail=str(exc)) from exc

    trade_uids = [_build_trade_uid(row) for row in validated_rows]
    existing = set(db.scalars(select(Trade.trade_uid).where(Trade.trade_uid.in_(trade_uids))).all())

    imports: list[Trade] = []
    duplicate_rows = 0
    for row, trade_uid in zip(validated_rows, trade_uids, strict=True):
        if trade_uid in existing:
            duplicate_rows += 1
            continue
        imports.append(
            Trade(
                trade_uid=trade_uid,
                account=row.account,
                symbol=row.symbol,
                trade_date=row.trade_date,
                side=row.side,
                quantity=row.quantity,
                price=row.price,
                fees=row.fees,
                currency=row.currency,
                broker_ref=row.broker_ref,
                source_file=file.filename,
                imported_at=datetime.now(UTC),
            )
        )

    if imports:
        db.add_all(imports)
        db.commit()

    return TradeImportResponse(
        filename=file.filename,
        total_rows=len(validated_rows),
        imported_rows=len(imports),
        duplicate_rows=duplicate_rows,
    )
